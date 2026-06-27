
import openpyxl
import os
import re


def get_instr_strs():

    value = os.environ.get('ROOT')
    assert value, "ERROR: ROOT varaiable must be set"
    wb = openpyxl.load_workbook(f'{value}/cpu/doc/commands.xlsx')
    ws = wb["Instruction_LUT"]  # or wb['Sheet1']

    instruction_lut = {}
    #instruction_lut["ADDI"] = "32'b0000x0000"

    # Read cell values
    for row in ws.iter_rows():
        # Skip header
        if(row[0].row == 1):
            continue
        
        
        instr_name = row[0].value
        assert instr_name not in instruction_lut.keys()


        bits = 0
        instr_str = ""
        
        for cell in row[1::]:
            feild = cell.value
            if(cell.value == None):
                break
            #print(feild)
            # Literal bits
            x = re.match("^[0|1]+$", feild)
            if(x):
                instr_str = x.group(0) + instr_str
                bits += len(x.group(0))
                continue

            # Register addr
            x = re.match("^(r(s1|s2|d)|shamt)$", feild)
            if(x):
                instr_str = "xxxxx" + instr_str
                bits += 5
                continue
            
            # Immediate
            x = re.match(f"^imm\[(.+)\]$", feild)
            if(x):
                immd_bits = 0
                for b in x.group(1).split("|"):
                    x = re.match("^(\d+)(:\d+)?$", b)
                    assert x, "ERROR: Bad immediate formating " + feild

                    if(not x.group(2)):
                        immd_bits += 1
                    else:
                        immd_bits += abs(int(x.group(1)) - int(x.group(2)[1::])) + 1
                bits += immd_bits
                instr_str = "x"*immd_bits + instr_str
                continue
            
            # Fence mode
            x = re.match(f"^fm$", feild)
            if(x):
                bits += 4
                instr_str = "x000" + instr_str
                continue

            # Fence pred/succsor set
            x = re.match(f"^pred|succ$", feild)
            if(x):
                bits += 4
                instr_str = "xxxx" + instr_str
                continue

            assert False, f"ERROR: Could not match column to regex: {feild}"

        
        assert bits == 32, f"ERROR: instruction does not have correct amount of bits {str(instr_name)} : {str(bits)}"
        instruction_lut[instr_name] = "32'b" + instr_str

    return instruction_lut



assertion_define = '''
// This file was generated with $ROOT/cpu/dv/scrpits/generate_opcode_assertions.py

`define COMPARE_ALL_INSTR ('''

if __name__ == "__main__":

    ROOT = os.environ.get('ROOT')
    assert ROOT, "ERROR: ROOT varaiable must be set"
    wb = openpyxl.load_workbook(f'{ROOT}/cpu/doc/commands.xlsx')

    ws = wb["Instruction_LUT"]  # or wb['Sheet1']

    instruction_lut = get_instr_strs()
    
    for inst_name in instruction_lut:
        #print(instruction_lut[inst_name])
        assertion_define += f"(instruction[31:0] ==? {instruction_lut[inst_name]}) || \\\n    "

    with open(f'{ROOT}/cpu/dv/hdl/instruction_assertion.svh', 'w') as f:
        f.write(assertion_define[:-10] + ") \n")
